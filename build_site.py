"""Build The Matrix static site from the Excel source.

Usage:
  python build_site.py "The Matrix.xlsx"

Output:
  Writes data.js (and optionally all site files if you want to regenerate).
"""

import json, re, sys, datetime
from openpyxl import load_workbook

def normalize_text(x):
    if x is None:
        return ""
    return str(x).replace("\r\n", "\n").replace("\r", "\n").strip()

def classify_outcome(outcome_text):
    t = (outcome_text or "").lower()
    if "conditional" in t:
        return "Conditional"
    if "unworkable" in t:
        return "Unworkable"
    if "workable" in t:
        return "Workable"
    return "Update"

def slugify(s):
    s = re.sub(r'[^a-zA-Z0-9]+', '-', s.strip().lower())
    s = re.sub(r'-+', '-', s).strip('-')
    return s or "category"

def main(path):
    wb = load_workbook(path)
    ws = wb.active

    def v(r,c):
        return ws.cell(row=r, column=c).value

    today = datetime.date.today().isoformat()

    left_rows = []
    for r in range(7, ws.max_row+1):
        task = v(r,3)
        if task is None:
            continue
        left_rows.append({
            "task": task,
            "status": v(r,4),
            "outcome": v(r,5),
            "steps": v(r,6),
        })

    right_rows = []
    for r in range(7, ws.max_row+1):
        task = v(r,8)
        if task is None:
            continue
        right_rows.append({"task": task, "status": v(r,9)})

    categories = []

    cat1 = {
        "id": "assemble-case-tasks",
        "name": "Assemble Case Tasks",
        "description": "Decision flow for Case Assembly based on task name/status: outcome and required steps.",
        "updates": []
    }

    for row in left_rows:
        title = normalize_text(row["task"]).split("\n")[0].strip() or "Untitled"
        cat1["updates"].append({
            "title": title,
            "date": today,
            "status": classify_outcome(row["outcome"]),
            "tags": [("status: " + normalize_text(row["status"]))] if row["status"] else [],
            "next_action": normalize_text(row["steps"]),
            "body": normalize_text(row["task"]),
            "outcome": normalize_text(row["outcome"]),
        })

    categories.append(cat1)

    cat2_name = normalize_text(v(5,8)) or "Legal Tool - ESCA IC Request"
    seen = set()
    cat2 = {
        "id": slugify(cat2_name),
        "name": cat2_name,
        "description": "Tasks listed under the Legal Tool section in the provided spreadsheet.",
        "updates": []
    }
    for row in right_rows:
        key = (normalize_text(row["task"]), normalize_text(row["status"]))
        if key in seen:
            continue
        seen.add(key)
        cat2["updates"].append({
            "title": key[0] or "Untitled",
            "date": today,
            "status": "Update",
            "tags": [("status: " + key[1])] if key[1] else [],
            "next_action": "",
            "body": "",
            "outcome": "",
        })

    if cat2["updates"]:
        categories.append(cat2)

    data = {"categories": categories}
    with open("data.js", "w", encoding="utf-8") as f:
        f.write("window.MATRIX_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n")

    print("Wrote data.js with", len(categories), "categories.")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python build_site.py <xlsx_path>")
        raise SystemExit(2)
    main(sys.argv[1])
